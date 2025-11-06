"""
Phoenix Manufacturing – Boots Inventory Tool
-------------------------------------------
Purpose:
- Provide a simple, reliable desktop app (Tkinter) to manage inventory stored in Excel.
- Operators can search, add, and remove items; all changes are logged for traceability.
- Avoid duplicate parts by canonicalizing PART# (uppercase, remove punctuation/spaces).
- Keep the team's Excel workflow (no DB to learn) while adding safety and speed.

Key Ideas:
- Storage: Excel workbook (openpyxl). Main sheet + append-only Log sheet.
- Invariants: (PART#, BOX#) uniqueness; QTY is integer >= 0; canonical PART#.
- Safety: confirm dialog before mutating; refuse invalid quantities; handle Excel lock.
- Resilience: header row is detected, not assumed; schema is auto-created if missing.
"""

import re
import sys
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from openpyxl import load_workbook, Workbook

# === Paths (PY script version: looks next to this .py) ===
# Use script directory so the Excel file lives beside the .py (simple deployment).
APP_DIR = Path(__file__).parent
EXCEL_PATH = APP_DIR / "molds.xlsx"   # keep molds.xlsx beside the .py
SHEET_NAME = "Sheet1"                 # data sheet (source of truth)
LOG_SHEET = "Log"                     # append-only audit log

# === Helpers ===
def normalize(s: str) -> str:
    """
    Canonicalize PART# to prevent duplicates caused by formatting variations.
    - Uppercase everything.
    - Strip spaces, dashes, dots, underscores, slashes, and backslashes.
    Examples:
      'ab-12/3' -> 'AB123'
      '  a.b _ c ' -> 'ABC'
    """
    if not s:
        return ""
    s = s.upper()
    s = re.sub(r'[\s\-._/\\]', '', s)
    return s

def find_header_row(ws) -> Optional[int]:
    """
    Robustly detect the header row instead of assuming it's row 1.
    - Scans up to 100 rows to find a line that 'looks like' headers.
    - Tolerates variation: 'PART#' or 'NAME'; 'QTY' or 'QT'; 'BOX#'/'BOX'/'BOX NUMBER', etc.
    This makes the tool resilient to introductory title rows or extra empty lines.
    """
    max_scan = min(100, ws.max_row)
    targets = {"PART#", "QTY", "BOX", "BOX#", "BOX #", "BOX NUMBER"}
    for i in range(1, max_scan + 1):
        # Pull raw row values and uppercase/strip for robust comparisons
        vals = [(ws.cell(row=i, column=j).value or "") for j in range(1, ws.max_column + 1)]
        up = [str(v).strip().upper() for v in vals]
        # Heuristic: must have a part-like and qty-like, and some box-like token
        if ("PART#" in up or "NAME" in up) and ("QTY" in up or "QT" in up) and any(t in up for t in targets):
            return i
    return None

def ensure_workbook() -> Tuple[Workbook, int, Dict[str, int]]:
    """
    Ensure the Excel file, main sheet schema, and Log sheet exist and are usable.
    Responsibilities:
    - Create 'molds.xlsx' if missing with basic headers.
    - Find or create the data sheet and required columns (PART#, QTY, BOX#).
    - Create the Log sheet with defined columns if missing.
    - Return: (workbook, header_row_index, column_mapping_dict)
      where column_mapping_dict maps 'PART#'/'QTY'/'BOX#' to actual column numbers.

    Design choice:
    - Column mapping is dynamic so we don't hardcode column indices.
    - If headers are missing, we add them to keep the sheet valid.
    """
    # Create a new workbook with minimal schema if file doesn't exist yet
    if not EXCEL_PATH.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        # Friendly title row + blank row + headers (makes human viewing nicer)
        ws.append(["STOCK ROOM BOOTS", "", datetime.now().date()])
        ws.append([])
        ws.append(["PART#", "QTY", "BOX#"])
        wb.save(EXCEL_PATH)

    # Open the workbook now that we are sure it exists
    wb = load_workbook(EXCEL_PATH)

    # Ensure main sheet exists with headers
    if SHEET_NAME not in wb.sheetnames:
        wb.create_sheet(SHEET_NAME)
        wb[SHEET_NAME].append(["PART#", "QTY", "BOX#"])
        wb.save(EXCEL_PATH)

    ws = wb[SHEET_NAME]
    # If header row isn't found, default to row 3 (matches our initial template)
    header_row = find_header_row(ws) or 3

    # Map headers to column indices (dynamic to tolerate column order/insertions)
    header_vals = [str(ws.cell(row=header_row, column=j).value or "").strip().upper()
                   for j in range(1, ws.max_column + 1)]
    col_map: Dict[str, int] = {}
    for j, name in enumerate(header_vals, start=1):
        if name in ("PART#", "NAME"):  # allow 'NAME' as legacy synonym
            col_map["PART#"] = j
        elif name in ("QTY", "QT"):
            col_map["QTY"] = j
        elif name in ("BOX#", "BOX", "BOX #", "BOX NUMBER"):
            col_map["BOX#"] = j

    # Ensure required columns exist (add any missing)
    for need in ("PART#", "QTY", "BOX#"):
        if need not in col_map:
            new_col = ws.max_column + 1
            ws.cell(row=header_row, column=new_col, value=need)
            col_map[need] = new_col

    # Ensure Log sheet exists with a fixed schema (append-only history)
    if LOG_SHEET not in wb.sheetnames:
        log = wb.create_sheet(LOG_SHEET)
        log.append(["Time", "User", "Action", "Part#", "QtyChange", "FromBox", "ToBox", "Notes"])

    return wb, header_row, col_map

def load_inventory(ws, header_row, col_map) -> List[Dict]:
    """
    Read the inventory rows into Python dicts while keeping the original row index.
    Returns a list of:
      { "row": excel_row_number, "PART#": str|None, "QTY": int, "BOX#": any }
    Behavior:
    - Skips fully empty lines and accidental repeated header rows.
    - Converts QTY safely to int; non-numeric -> 0 (defensive reading).
    - Keeps BOX# type flexible (could be int/str depending on Excel).
    """
    rows: List[Dict] = []
    for r in range(header_row + 1, ws.max_row + 1):
        part = ws.cell(row=r, column=col_map["PART#"]).value
        qty  = ws.cell(row=r, column=col_map["QTY"]).value
        box  = ws.cell(row=r, column=col_map["BOX#"]).value

        # Skip empty lines entirely
        if part is None and qty is None and box is None:
            continue

        # Skip stray header duplicates inside body
        if str(part).strip().upper() in ("PART#", "NAME"):
            continue

        # Safe qty parsing (junk -> 0)
        try:
            qty = int(qty) if qty is not None and str(qty).strip() != "" else 0
        except Exception:
            qty = 0

        rows.append({
            "row": r,
            "PART#": None if part is None else str(part).strip(),
            "QTY": qty,
            "BOX#": box
        })
    return rows

def write_log(wb, user: str, action: str, part: str, qty_delta=0, from_box="", to_box="", notes=""):
    """
    Append a single line to the Log sheet for full auditability.
    Columns: Time, User, Action, Part#, QtyChange, FromBox, ToBox, Notes
    Examples of Action:
      - 'ADD/INCR', 'ADD_NEW', 'REMOVE/DECR', 'REMOVE/DELETE_ROW', 'NORMALIZE_ALL'
    """
    log = wb[LOG_SHEET]
    log.append([
        datetime.now().isoformat(timespec="seconds"),
        user, action, part, qty_delta, from_box, to_box, notes
    ])

def normalize_all(user: str = "") -> int:
    """
    Normalize all PART# values in-place.
    - Rewrites PART# to canonical form to eliminate formatting duplicates.
    - Logs a single summary entry with count of changed rows.
    Returns: number of items normalized.
    """
    wb, header_row, col_map = ensure_workbook()
    ws = wb[SHEET_NAME]
    changed = 0

    # Iterate the column below the header; rewrite any non-canonical entries
    for r in range(header_row + 1, ws.max_row + 1):
        part = ws.cell(row=r, column=col_map["PART#"]).value
        if part:
            canon = normalize(str(part))
            if str(part) != canon:
                ws.cell(row=r, column=col_map["PART#"]).value = canon
                changed += 1

    # Only log/save if we actually modified something
    if changed:
        write_log(wb, user, "NORMALIZE_ALL", "", 0, "", "", f"{changed} items")
        wb.save(EXCEL_PATH)
    return changed

# === Inventory Ops (PART# + BOX# unique) ===
def add_mode(part_input: str, qty: int, box: Optional[int], user: str) -> str:
    """
    Add or increment inventory entries.
    Rules:
      - If (PART#, BOX#) exists -> increment its qty.
      - If same PART# exists but different BOX# specified -> create a NEW row for that box.
      - If multiple boxes exist for that PART# and box is not specified -> ask for a box (avoid guessing).
      - If exactly one row exists and no box specified -> increment that one (good UX).
      - Always write a log entry and persist changes.

    Returns a human-readable message for UI dialogs.
    """
    wb, header_row, col_map = ensure_workbook()
    ws = wb[SHEET_NAME]
    items = load_inventory(ws, header_row, col_map)

    # Canonicalize input so comparisons are consistent
    canon = normalize(part_input)
    matches = [it for it in items if normalize(it["PART#"]) == canon]

    # Ambiguity: multiple boxes exist but no specific box given
    if len(matches) > 1 and box is None:
        return "Multiple boxes exist for this part. Please enter Box#."

    # Convenience path: exactly one row exists; increment it if no box given
    if len(matches) == 1 and box is None:
        match = matches[0]
        new_qty = match["QTY"] + qty
        ws.cell(row=match["row"], column=col_map["QTY"]).value = new_qty
        ws.cell(row=match["row"], column=col_map["PART#"]).value = canon  # enforce canonical
        write_log(wb, user, "ADD/INCR", canon, qty_delta=qty, to_box=match["BOX#"])
        wb.save(EXCEL_PATH)
        return f"Updated {canon}: {match['QTY']} → {new_qty}  (Box: {match['BOX#']})"

    # If a box was specified, try to find that exact (PART#, BOX#) row
    if box is not None:
        same_box = next((it for it in matches if (it["BOX#"] == box)), None)
        if same_box:
            # Found exact row -> increment
            new_qty = same_box["QTY"] + qty
            ws.cell(row=same_box["row"], column=col_map["QTY"]).value = new_qty
            ws.cell(row=same_box["row"], column=col_map["PART#"]).value = canon
            write_log(wb, user, "ADD/INCR", canon, qty_delta=qty, to_box=box)
            wb.save(EXCEL_PATH)
            return f"Updated {canon}: {same_box['QTY']} → {new_qty}  (Box: {box})"
        else:
            # Create a brand new row for this (PART#, BOX#)
            last_row = ws.max_row + 1
            ws.cell(row=last_row, column=col_map["PART#"]).value = canon
            ws.cell(row=last_row, column=col_map["QTY"]).value = qty
            ws.cell(row=last_row, column=col_map["BOX#"]).value = box
            write_log(wb, user, "ADD_NEW", canon, qty_delta=qty, to_box=box, notes="new box")
            wb.save(EXCEL_PATH)
            return f"Added {canon}  (qty {qty}, box {box})."

    # No matches at all and no box provided -> we need Box# to create the first row
    if not matches and box is None:
        return "No existing row for this part. Please enter Box# to create a new one."

    # Fallback: create a new row (covers not-matches + provided box)
    last_row = ws.max_row + 1
    ws.cell(row=last_row, column=col_map["PART#"]).value = canon
    ws.cell(row=last_row, column=col_map["QTY"]).value = qty
    ws.cell(row=last_row, column=col_map["BOX#"]).value = box
    write_log(wb, user, "ADD_NEW", canon, qty_delta=qty, to_box=box, notes="created")
    wb.save(EXCEL_PATH)
    return f"Added {canon}  (qty {qty}, box {box})."

def remove_mode(part_input: str, qty: int, user: str, box: Optional[int] = None) -> str:
    """
    Remove/decrement inventory entries.
    Rules:
      - If box is specified -> operate on that exact (PART#, BOX#) row.
      - If multiple boxes exist and no box specified -> ask for a box (avoid wrong row).
      - If only one row exists for that part -> operate on it directly.
      - If qty would go negative -> refuse with an error.
      - If qty hits zero -> delete the row to keep the sheet clean.
      - Always log the change and persist.

    Returns a human-readable message for UI dialogs.
    """
    wb, header_row, col_map = ensure_workbook()
    ws = wb[SHEET_NAME]
    items = load_inventory(ws, header_row, col_map)

    # Canonicalize to find target rows reliably
    canon = normalize(part_input)
    matches = [it for it in items if normalize(it["PART#"]) == canon]

    # Part doesn't exist at all
    if not matches:
        return "Code not found."

    # Basic input guard
    if qty <= 0:
        return "Quantity must be positive."

    # Choose the exact row to operate on
    if box is not None:
        target = next((it for it in matches if it["BOX#"] == box), None)
        if not target:
            return f"No row found for {canon} in Box {box}."
    else:
        if len(matches) > 1:
            return "Multiple boxes exist for this part. Please enter Box#."
        target = matches[0]

    # Validate that we don't remove more than available
    current_qty = target["QTY"]
    if current_qty - qty < 0:
        return f"Cannot remove {qty}; only {current_qty} available."

    new_qty = current_qty - qty

    if new_qty == 0:
        # If depleting to zero, delete the entire row (and log the deletion)
        write_log(
            wb, user, "REMOVE/DELETE_ROW", canon,
            qty_delta=-qty, from_box=target["BOX#"],
            notes="qty hit 0 -> row deleted"
        )
        ws.delete_rows(target["row"], 1)
        wb.save(EXCEL_PATH)
        return f"Removed {qty} from {canon} (Box: {target['BOX#']}); qty is 0, row deleted."
    else:
        # Decrement and keep the row
        ws.cell(row=target["row"], column=col_map["PART#"]).value = canon  # enforce canonical
        ws.cell(row=target["row"], column=col_map["QTY"]).value = new_qty
        write_log(wb, user, "REMOVE/DECR", canon, qty_delta=-qty, from_box=target["BOX#"])
        wb.save(EXCEL_PATH)
        return f"Updated {canon}: {current_qty} → {new_qty} (Box: {target['BOX#']})"

# === Confirmation Dialog (green for ADD, red for REMOVE) ===
def confirm_action(root: tk.Tk, action: str, part: str, qty: int, box: Optional[int]) -> bool:
    """
    Modal confirmation dialog to reduce misclicks.
    - Green title for ADD and red for REMOVE to reinforce caution.
    - Centers over the main window.
    - Returns True if user confirms, False otherwise.
    """
    win = tk.Toplevel(root)
    win.title("Confirm action")
    win.transient(root)    # keep on top of the parent
    win.grab_set()         # modal behavior
    win.resizable(False, False)

    is_add = (action.upper() == "ADD")
    color = "#0B8A00" if is_add else "#C1121F"

    wrap = ttk.Frame(win, padding=16)
    wrap.pack(fill="both", expand=True)

    tk.Label(
        wrap,
        text="Confirm " + action.upper(),
        font=("Segoe UI", 14, "bold"),
        fg=color
    ).pack(anchor="w")

    # Show normalized part and provided params for clarity
    canon = normalize(part or "")
    details = "Part#: {0}\nQty: {1}".format(canon, qty)
    if box is not None:
        details += "\nBox#: {0}".format(box)
    tk.Label(wrap, text=details, font=("Segoe UI", 11)).pack(anchor="w", pady=(8, 6))

    # Buttons row
    btns = ttk.Frame(wrap)
    btns.pack(anchor="e", pady=(8, 0))
    result = {"ok": False}

    def _ok():
        result["ok"] = True
        win.destroy()

    def _cancel():
        result["ok"] = False
        win.destroy()

    ttk.Button(btns, text="Cancel", command=_cancel).grid(row=0, column=0, padx=(0, 8))
    ttk.Button(btns, text="Confirm", command=_ok).grid(row=0, column=1)

    # Center dialog over the parent window
    root.update_idletasks()
    x = root.winfo_rootx() + (root.winfo_width() // 2) - 160
    y = root.winfo_rooty() + (root.winfo_height() // 2) - 80
    win.geometry("+{0}+{1}".format(x, y))

    win.wait_window()
    return result["ok"]

# === GUI ===
class App(tk.Tk):
    """
    Tkinter main window:
    - Top panel: user name (for logging), mode toggle (Add/Remove), live search, inputs, and actions.
    - Middle: Treeview table (PART#, QTY, BOX#) with scrollbar.
    - Status line: short feedback.
    - Behavior: selecting a row pre-fills the form; search filters by canonical substring.
    """
    def __init__(self):
        super().__init__()
        self.title("Boots Inventory (Add / Remove)")
        self.geometry("960x600")

        # UI state variables (StringVar for Tk binding/updates)
        self.user_var = tk.StringVar()              # used in logs
        self.mode_var = tk.StringVar(value="ADD")   # "ADD" or "REMOVE"
        self.search_var = tk.StringVar()            # live filter
        self.code_var = tk.StringVar()              # input PART#
        self.qty_var = tk.StringVar(value="1")      # input quantity (as text; parsed later)
        self.box_var = tk.StringVar()               # input box (optional)
        self.msg_var = tk.StringVar(value="Ready.") # status message

        # Build UI and initial table load
        self._build_top()
        self._build_table()
        self._bind_events()
        self.refresh_table()

    def _build_top(self):
        """
        Top control strip:
        - User name (appears in Log sheet).
        - Mode toggle (Add/Remove).
        - Search box (live).
        - Entry fields for Part#, Qty, Box#.
        - Action buttons: Apply, Normalize All, Open in Excel.
        """
        top = ttk.Frame(self, padding=10)
        top.pack(fill="x")

        ttk.Label(top, text="User:").grid(row=0, column=0, sticky="e", padx=(0, 6))
        ttk.Entry(top, textvariable=self.user_var, width=14).grid(row=0, column=1, sticky="w", padx=(0, 10))

        # Small readout to mirror the user (nice feedback in demos)
        ttk.Label(top, text="Added by:").grid(row=0, column=2, sticky="e", padx=(8, 6))
        ttk.Label(top, textvariable=self.user_var).grid(row=0, column=3, sticky="w")

        # Mode toggle buttons
        ttk.Radiobutton(top, text="Add mode", value="ADD",    variable=self.mode_var).grid(row=0, column=4, sticky="w", padx=(16, 0))
        ttk.Radiobutton(top, text="Remove mode", value="REMOVE", variable=self.mode_var).grid(row=0, column=5, sticky="w")

        # Live search (filters table immediately)
        ttk.Label(top, text="Search:").grid(row=0, column=6, sticky="e", padx=(16, 6))
        ttk.Entry(top, textvariable=self.search_var, width=28).grid(row=0, column=7, sticky="w")

        ttk.Button(top, text="Refresh", command=self.refresh_table).grid(row=0, column=8, sticky="w", padx=(10, 0))

        # Input fields row
        ttk.Label(top, text="Part#:").grid(row=1, column=0, sticky="e", padx=(0, 6), pady=(10, 0))
        ttk.Entry(top, textvariable=self.code_var, width=28).grid(row=1, column=1, columnspan=2, sticky="w", pady=(10, 0))

        ttk.Label(top, text="Qty:").grid(row=1, column=3, sticky="e", padx=(10, 6), pady=(10, 0))
        ttk.Entry(top, textvariable=self.qty_var, width=8).grid(row=1, column=4, sticky="w", pady=(10, 0))

        ttk.Label(top, text="Box# (optional):").grid(row=1, column=5, sticky="e", padx=(10, 6), pady=(10, 0))
        ttk.Entry(top, textvariable=self.box_var, width=10).grid(row=1, column=6, sticky="w", pady=(10, 0))

        # Buttons row
        btns = ttk.Frame(top)
        btns.grid(row=2, column=0, columnspan=9, sticky="w", pady=(12, 0))
        ttk.Button(btns, text="Apply",          command=self.on_apply).grid(row=0, column=0, padx=(0, 10))
        ttk.Button(btns, text="Normalize All",  command=self.on_normalize_all).grid(row=0, column=1, padx=(0, 10))
        ttk.Button(btns, text="Open in Excel",  command=self.open_excel).grid(row=0, column=2, padx=(0, 10))

        # Status line
        status = ttk.Frame(self, padding=(10, 0, 10, 10))
        status.pack(fill="x")
        ttk.Label(status, textvariable=self.msg_var).pack(side="left")

        # Make the grid responsive so fields expand with window
        for c in range(9):
            top.grid_columnconfigure(c, weight=1)

    def _build_table(self):
        """
        Middle table area (Treeview):
        - Shows PART#, QTY, BOX#.
        - Scrollbar for long lists.
        - Columns sized for readability.
        """
        mid = ttk.Frame(self, padding=(10, 0, 10, 10))
        mid.pack(fill="both", expand=True)

        cols = ("PART#", "QTY", "BOX#")
        self.tree = ttk.Treeview(mid, columns=cols, show="headings", selectmode="browse")
        for c in cols:
            self.tree.heading(c, text=c)
        self.tree.column("PART#", anchor="w", width=360)
        self.tree.column("QTY", anchor="center", width=100)
        self.tree.column("BOX#", anchor="center", width=100)
        self.tree.pack(side="left", fill="both", expand=True)

        # Vertical scrollbar
        vsb = ttk.Scrollbar(mid, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")

    def _bind_events(self):
        """
        Event bindings:
        - Search box triggers table refresh (live filtering).
        - Selecting a row fills the input fields (saves typing).
        """
        self.search_var.trace_add("write", lambda *_: self.refresh_table())
        self.tree.bind("<<TreeviewSelect>>", self.on_select_row)

    def on_select_row(self, _evt=None):
        """
        When the user clicks a table row:
        - Pre-fill 'Part#' with that row's part.
        - Set Qty to '1' (common quick action).
        - Pre-fill Box# if present.
        """
        sel = self.tree.selection()
        if not sel:
            return
        vals = self.tree.item(sel[0], "values")
        if not vals:
            return
        part, qty, box = vals
        self.code_var.set(part)
        self.qty_var.set("1")
        self.box_var.set("" if box is None else str(box))

    def _parse_int(self, s: str, allow_none=False) -> Optional[int]:
        """
        Safe integer parsing for text entries.
        - Returns None if blank and allow_none=True.
        - Returns None on any parsing error to trigger validation messages.
        """
        s = (s or "").strip()
        if s == "" and allow_none:
            return None
        try:
            return int(s)
        except Exception:
            return None

    def refresh_table(self):
        """
        Reload the table from the workbook:
        - Reads items via load_inventory.
        - Applies canonical substring filter from the search box.
        - Sorts by PART#, then BOX#.
        - Updates status line with item count.
        """
        # Clear existing rows
        for iid in self.tree.get_children():
            self.tree.delete(iid)

        wb, header_row, col_map = ensure_workbook()
        ws = wb[SHEET_NAME]
        items = load_inventory(ws, header_row, col_map)

        # Live filter: compare normalized forms for consistency
        term = normalize(self.search_var.get())
        if term:
            items = [it for it in items if term in normalize(it["PART#"] or "")]

        # Sort for a stable, predictable view
        items.sort(key=lambda it: (str(it["PART#"] or ""), it["BOX#"] or 0))
        for it in items:
            self.tree.insert("", "end", values=(it["PART#"], it["QTY"], it["BOX#"]))

        self.msg_var.set("{0} items shown.".format(len(items)))

    def on_apply(self):
        """
        Handle the Apply button:
        - Validate inputs (Part#, positive Qty).
        - Confirm action with a color-coded modal.
        - Route to add_mode/remove_mode.
        - Show context-appropriate dialogs and refresh the table.
        - Handle PermissionError if Excel is open/locked.
        """
        user = self.user_var.get().strip()
        code = self.code_var.get().strip()
        qty = self._parse_int(self.qty_var.get())
        box = self._parse_int(self.box_var.get(), allow_none=True)

        # Basic validation before touching the file
        if not code or qty is None or qty <= 0:
            messagebox.showerror("Input error", "Enter a Part# and a positive Qty.")
            return

        try:
            # Confirm action before mutating
            if self.mode_var.get() == "ADD":
                if not confirm_action(self, "ADD", code, qty, box):
                    self.msg_var.set("Add cancelled.")
                    return
                msg = add_mode(code, qty, box, user)
            else:
                if not confirm_action(self, "REMOVE", code, qty, box):
                    self.msg_var.set("Remove cancelled.")
                    return
                msg = remove_mode(code, qty, user, box=box)

            # Feedback depending on the message
            if msg == "Code not found." or msg.startswith("Multiple boxes exist") or msg.startswith("No existing row"):
                messagebox.showwarning("Notice", msg)
            elif msg.startswith("Cannot remove"):
                messagebox.showerror("Error", msg)
            else:
                messagebox.showinfo("Done", msg)

            self.msg_var.set(msg)
            self.refresh_table()

        except PermissionError:
            # Common in Excel workflows when file is open elsewhere
            messagebox.showerror("Locked", "Close the Excel file and try again.")
        except Exception as e:
            # Defensive catch-all for unexpected errors
            messagebox.showerror("Error", str(e))

    def on_normalize_all(self):
        """
        Trigger normalization across all PART# entries.
        - Good for cleaning legacy rows.
        - Logs a summary entry with how many rows were changed.
        """
        try:
            changed = normalize_all(self.user_var.get().strip())
            self.msg_var.set("Normalized {0} item(s).".format(changed))
            self.refresh_table()
            if changed:
                messagebox.showinfo("Normalization complete", "Normalized {0} item(s).".format(changed))
        except PermissionError:
            messagebox.showerror("Locked", "Close the Excel file and try again.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def open_excel(self):
        """
        Convenience: open the Excel file in the OS default handler.
        - Windows: os.startfile
        - macOS:   open
        - Linux:   xdg-open
        Useful for quick manual checks or printing.
        """
        try:
            path = EXCEL_PATH.resolve()
            import os, platform, subprocess
            if platform.system() == "Windows":
                os.startfile(str(path))
            elif platform.system() == "Darwin":
                subprocess.run(["open", str(path)])
            else:
                subprocess.run(["xdg-open", str(path)])
        except Exception as e:
            messagebox.showerror("Open failed", str(e))

# Entry point: create and run the Tkinter app.
if __name__ == "__main__":
    app = App()
    app.mainloop()
