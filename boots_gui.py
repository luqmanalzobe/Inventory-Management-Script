import re
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from openpyxl import load_workbook, Workbook

# === Configuration ===
EXCEL_PATH = Path("molds.xlsx")   # <- change to your Excel filename if needed
SHEET_NAME = "Sheet1"
LOG_SHEET = "Log"

# === Helpers ===
def normalize(s: str) -> str:
    """UPPERCASE and remove spaces, dashes, dots, underscores, and slashes."""
    if not s:
        return ""
    s = s.upper()
    s = re.sub(r'[\s\-._/\\]', '', s)
    return s

def find_header_row(ws) -> Optional[int]:
    max_scan = min(100, ws.max_row)
    targets = {"PART#", "QTY", "BOX", "BOX#", "BOX #", "BOX NUMBER"}
    for i in range(1, max_scan + 1):
        vals = [(ws.cell(row=i, column=j).value or "") for j in range(1, ws.max_column + 1)]
        up = [str(v).strip().upper() for v in vals]
        if ("PART#" in up or "NAME" in up) and ("QTY" in up or "QT" in up) and any(t in up for t in targets):
            return i
    return None

def ensure_workbook() -> Tuple[Workbook, int, Dict[str, int]]:
    """Ensure workbook, sheet, headers, and Log sheet exist."""
    if not EXCEL_PATH.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(["STOCK ROOM BOOTS", "", datetime.now().date()])
        ws.append([])
        ws.append(["PART#", "QTY", "BOX#"])
        wb.save(EXCEL_PATH)

    wb = load_workbook(EXCEL_PATH)

    if SHEET_NAME not in wb.sheetnames:
        wb.create_sheet(SHEET_NAME)
        wb[SHEET_NAME].append(["PART#", "QTY", "BOX#"])
        wb.save(EXCEL_PATH)

    ws = wb[SHEET_NAME]
    header_row = find_header_row(ws) or 3

    # Map headers to columns
    header_vals = [str(ws.cell(row=header_row, column=j).value or "").strip().upper()
                   for j in range(1, ws.max_column + 1)]
    col_map: Dict[str, int] = {}
    for j, name in enumerate(header_vals, start=1):
        if name in ("PART#", "NAME"):
            col_map["PART#"] = j
        elif name in ("QTY", "QT"):
            col_map["QTY"] = j
        elif name in ("BOX#", "BOX", "BOX #", "BOX NUMBER"):
            col_map["BOX#"] = j

    # Ensure required columns
    for need in ("PART#", "QTY", "BOX#"):
        if need not in col_map:
            new_col = ws.max_column + 1
            ws.cell(row=header_row, column=new_col, value=need)
            col_map[need] = new_col

    # Ensure Log
    if LOG_SHEET not in wb.sheetnames:
        log = wb.create_sheet(LOG_SHEET)
        log.append(["Time", "User", "Action", "Part#", "QtyChange", "FromBox", "ToBox", "Notes"])

    return wb, header_row, col_map

def load_inventory(ws, header_row, col_map) -> List[Dict]:
    rows: List[Dict] = []
    for r in range(header_row + 1, ws.max_row + 1):
        part = ws.cell(row=r, column=col_map["PART#"]).value
        qty  = ws.cell(row=r, column=col_map["QTY"]).value
        box  = ws.cell(row=r, column=col_map["BOX#"]).value
        if part is None and qty is None and box is None:
            continue
        if str(part).strip().upper() in ("PART#", "NAME"):
            continue
        try:
            qty = int(qty) if qty is not None and str(qty).strip() != "" else 0
        except Exception:
            qty = 0
        rows.append({"row": r, "PART#": None if part is None else str(part).strip(), "QTY": qty, "BOX#": box})
    return rows

def write_log(wb, user: str, action: str, part: str, qty_delta=0, from_box="", to_box="", notes=""):
    log = wb[LOG_SHEET]
    log.append([datetime.now().isoformat(timespec="seconds"), user, action, part, qty_delta, from_box, to_box, notes])

def normalize_all(user: str = "") -> int:
    wb, header_row, col_map = ensure_workbook()
    ws = wb[SHEET_NAME]
    changed = 0
    for r in range(header_row + 1, ws.max_row + 1):
        part = ws.cell(row=r, column=col_map["PART#"]).value
        if part:
            canon = normalize(str(part))
            if str(part) != canon:
                ws.cell(row=r, column=col_map["PART#"]).value = canon
                changed += 1
    if changed:
        write_log(wb, user, "NORMALIZE_ALL", "", 0, "", "", f"{changed} items")
        wb.save(EXCEL_PATH)
    return changed

# === Inventory Ops (treat PART# + BOX# as unique) ===
def add_mode(part_input: str, qty: int, box: Optional[int], user: str) -> str:
    """
    If same PART#+BOX# exists -> increment that row.
    If same PART# but different BOX# -> create a NEW row (do NOT modify the old box).
    If multiple boxes exist for that part and no box provided -> ask for box.
    """
    wb, header_row, col_map = ensure_workbook()
    ws = wb[SHEET_NAME]
    items = load_inventory(ws, header_row, col_map)

    canon = normalize(part_input)
    matches = [it for it in items if normalize(it["PART#"]) == canon]

    # If multiple boxes exist and no box provided, require a box
    if len(matches) > 1 and box is None:
        return "Multiple boxes exist for this part. Please enter Box#."

    # If exactly one match and no box provided, increment that row
    if len(matches) == 1 and box is None:
        match = matches[0]
        new_qty = match["QTY"] + qty
        ws.cell(row=match["row"], column=col_map["QTY"]).value = new_qty
        ws.cell(row=match["row"], column=col_map["PART#"]).value = canon
        write_log(wb, user, "ADD/INCR", canon, qty_delta=qty, to_box=match["BOX#"])
        wb.save(EXCEL_PATH)
        return f"Updated {canon}: {match['QTY']} → {new_qty}  (Box: {match['BOX#']})"

    # If a box is provided, check for same-part same-box
    if box is not None:
        same_box = next((it for it in matches if (it["BOX#"] == box)), None)
        if same_box:
            # increment that row
            new_qty = same_box["QTY"] + qty
            ws.cell(row=same_box["row"], column=col_map["QTY"]).value = new_qty
            ws.cell(row=same_box["row"], column=col_map["PART#"]).value = canon
            write_log(wb, user, "ADD/INCR", canon, qty_delta=qty, to_box=box)
            wb.save(EXCEL_PATH)
            return f"Updated {canon}: {same_box['QTY']} → {new_qty}  (Box: {box})"
        else:
            # create new row for this new box
            last_row = ws.max_row + 1
            ws.cell(row=last_row, column=col_map["PART#"]).value = canon
            ws.cell(row=last_row, column=col_map["QTY"]).value = qty
            ws.cell(row=last_row, column=col_map["BOX#"]).value = box
            write_log(wb, user, "ADD_NEW", canon, qty_delta=qty, to_box=box, notes="new box")
            wb.save(EXCEL_PATH)
            return f"Added {canon}  (qty {qty}, box {box})."

    # No match and no box specified -> require box to create a new line
    if not matches and box is None:
        return "No existing row for this part. Please enter Box# to create a new one."

    # Fallback: create new row (part existed in other boxes or none at all)
    last_row = ws.max_row + 1
    ws.cell(row=last_row, column=col_map["PART#"]).value = canon
    ws.cell(row=last_row, column=col_map["QTY"]).value = qty
    ws.cell(row=last_row, column=col_map["BOX#"]).value = box
    write_log(wb, user, "ADD_NEW", canon, qty_delta=qty, to_box=box, notes="created")
    wb.save(EXCEL_PATH)
    return f"Added {canon}  (qty {qty}, box {box})."

def remove_mode(part_input: str, qty: int, user: str, box: Optional[int] = None) -> str:
    """
    If BOX# is provided -> remove from that exact PART# + BOX# row.
    If multiple boxes exist and BOX# not provided -> ask for box.
    If exactly one matching row -> remove from it.
    """
    wb, header_row, col_map = ensure_workbook()
    ws = wb[SHEET_NAME]
    items = load_inventory(ws, header_row, col_map)

    canon = normalize(part_input)
    matches = [it for it in items if normalize(it["PART#"]) == canon]

    if not matches:
        return "Code not found."
    if qty <= 0:
        return "Quantity must be positive."

    # If box given, target that row
    if box is not None:
        target = next((it for it in matches if it["BOX#"] == box), None)
        if not target:
            return f"No row found for {canon} in Box {box}."
    else:
        # No box: if multiple boxes, require a box to avoid ambiguity
        if len(matches) > 1:
            return "Multiple boxes exist for this part. Please enter Box#."
        target = matches[0]

    if target["QTY"] - qty < 0:
        return f"Cannot remove {qty}; only {target['QTY']} available."

    new_qty = target["QTY"] - qty
    ws.cell(row=target["row"], column=col_map["PART#"]).value = canon  # enforce canonical
    ws.cell(row=target["row"], column=col_map["QTY"]).value = new_qty
    write_log(wb, user, "REMOVE/DECR", canon, qty_delta=-qty, from_box=target["BOX#"])
    wb.save(EXCEL_PATH)
    return f"Updated {canon}: {target['QTY']} → {new_qty} (Box: {target['BOX#']})"

# === Confirmation Dialog ===
def confirm_action(root: tk.Tk, action: str, part: str, qty: int, box: Optional[int]) -> bool:
    """
    Custom modal confirmation.
      action: "ADD" or "REMOVE"
      returns True if user confirms
    """
    win = tk.Toplevel(root)
    win.title("Confirm action")
    win.transient(root)
    win.grab_set()  # modal
    win.resizable(False, False)

    # Colors & text
    is_add = (action.upper() == "ADD")
    color = "#0B8A00" if is_add else "#C1121F"   # green vs red
    big = ttk.Frame(win, padding=16)
    big.pack(fill="both", expand=True)

    # Bold label
    title = f"Confirm {action.upper()}"
    lbl = tk.Label(big, text=title, font=("Segoe UI", 14, "bold"), fg=color)
    lbl.pack(anchor="w")

    canon = normalize(part or "")
    details = f"Part#: {canon}\nQty: {qty}"
    if box is not None:
        details += f"\nBox#: {box}"
    tk.Label(big, text=details, font=("Segoe UI", 11)).pack(anchor="w", pady=(8, 6))

    # Buttons
    btns = ttk.Frame(big)
    btns.pack(anchor="e", pady=(8, 0))
    result = {"ok": False}

    def _ok():
        result["ok"] = True
        win.destroy()

    def _cancel():
        result["ok"] = False
        win.destroy()

    ttk.Button(btns, text="Cancel", command=_cancel).grid(row=0, column=0, padx=(0, 8))
    ttk.Button(btns, text="Confirm", command=_ok).grid(row=0, column=1)

    # Center over parent
    root.update_idletasks()
    x = root.winfo_rootx() + (root.winfo_width() // 2) - 160
    y = root.winfo_rooty() + (root.winfo_height() // 2) - 80
    win.geometry(f"+{x}+{y}")

    win.wait_window()
    return result["ok"]

# === GUI ===
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Boots Inventory (Add / Remove)")
        self.geometry("960x600")

        self.user_var = tk.StringVar()
        self.mode_var = tk.StringVar(value="ADD")
        self.search_var = tk.StringVar()
        self.code_var = tk.StringVar()
        self.qty_var = tk.StringVar(value="1")
        self.box_var = tk.StringVar()
        self.msg_var = tk.StringVar(value="Ready.")

        self._build_top()
        self._build_table()
        self._bind_events()
        self.refresh_table()

    def _build_top(self):
        top = ttk.Frame(self, padding=10)
        top.pack(fill="x")

        # Row 0
        ttk.Label(top, text="User:").grid(row=0, column=0, sticky="e", padx=(0, 6))
        ttk.Entry(top, textvariable=self.user_var, width=14).grid(row=0, column=1, sticky="w", padx=(0, 10))

        ttk.Label(top, text="Added by:").grid(row=0, column=2, sticky="e", padx=(8, 6))
        ttk.Label(top, textvariable=self.user_var).grid(row=0, column=3, sticky="w")

        ttk.Radiobutton(top, text="Add mode", value="ADD", variable=self.mode_var).grid(row=0, column=4, sticky="w", padx=(16, 0))
        ttk.Radiobutton(top, text="Remove mode", value="REMOVE", variable=self.mode_var).grid(row=0, column=5, sticky="w")

        ttk.Label(top, text="Search:").grid(row=0, column=6, sticky="e", padx=(16, 6))
        ttk.Entry(top, textvariable=self.search_var, width=28).grid(row=0, column=7, sticky="w")

        ttk.Button(top, text="Refresh", command=self.refresh_table).grid(row=0, column=8, sticky="w", padx=(10, 0))

        # Row 1
        ttk.Label(top, text="Part#:").grid(row=1, column=0, sticky="e", padx=(0, 6), pady=(10, 0))
        ttk.Entry(top, textvariable=self.code_var, width=28).grid(row=1, column=1, columnspan=2, sticky="w", pady=(10, 0))

        ttk.Label(top, text="Qty:").grid(row=1, column=3, sticky="e", padx=(10, 6), pady=(10, 0))
        ttk.Entry(top, textvariable=self.qty_var, width=8).grid(row=1, column=4, sticky="w", pady=(10, 0))

        ttk.Label(top, text="Box# (optional):").grid(row=1, column=5, sticky="e", padx=(10, 6), pady=(10, 0))
        ttk.Entry(top, textvariable=self.box_var, width=10).grid(row=1, column=6, sticky="w", pady=(10, 0))

        # Row 2 - Buttons
        btns = ttk.Frame(top)
        btns.grid(row=2, column=0, columnspan=9, sticky="w", pady=(12, 0))
        ttk.Button(btns, text="Apply", command=self.on_apply).grid(row=0, column=0, padx=(0, 10))
        ttk.Button(btns, text="Normalize All", command=self.on_normalize_all).grid(row=0, column=1, padx=(0, 10))
        ttk.Button(btns, text="Open in Excel", command=self.open_excel).grid(row=0, column=2, padx=(0, 10))

        # Status
        status = ttk.Frame(self, padding=(10, 0, 10, 10))
        status.pack(fill="x")
        ttk.Label(status, textvariable=self.msg_var).pack(side="left")

        for c in range(9):
            top.grid_columnconfigure(c, weight=1)

    def _build_table(self):
        mid = ttk.Frame(self, padding=(10, 0, 10, 10))
        mid.pack(fill="both", expand=True)

        cols = ("PART#", "QTY", "BOX#")
        self.tree = ttk.Treeview(mid, columns=cols, show="headings", selectmode="browse")
        for c in cols:
            self.tree.heading(c, text=c)
            anchor = "w" if c == "PART#" else "center"
            width = 360 if c == "PART#" else 100
            self.tree.column(c, anchor=anchor, width=width)
        self.tree.pack(side="left", fill="both", expand=True)

        vsb = ttk.Scrollbar(mid, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")

    def _bind_events(self):
        self.search_var.trace_add("write", lambda *_: self.refresh_table())
        self.tree.bind("<<TreeviewSelect>>", self.on_select_row)

    def on_select_row(self, _evt=None):
        sel = self.tree.selection()
        if not sel:
            return
        vals = self.tree.item(sel[0], "values")
        if not vals:
            return
        part, qty, box = vals
        self.code_var.set(part)
        self.qty_var.set("1")
        self.box_var.set("" if box is None else str(box))

    def _parse_int(self, s: str, allow_none=False) -> Optional[int]:
        s = (s or "").strip()
        if s == "" and allow_none:
            return None
        try:
            return int(s)
        except Exception:
            return None

    def refresh_table(self):
        for iid in self.tree.get_children():
            self.tree.delete(iid)

        wb, header_row, col_map = ensure_workbook()
        ws = wb[SHEET_NAME]
        items = load_inventory(ws, header_row, col_map)

        term = normalize(self.search_var.get())
        if term:
            items = [it for it in items if term in normalize(it["PART#"] or "")]

        # sort by part then box
        items.sort(key=lambda it: (str(it["PART#"] or ""), it["BOX#"] or 0))
        for it in items:
            self.tree.insert("", "end", values=(it["PART#"], it["QTY"], it["BOX#"]))

        self.msg_var.set(f"{len(items)} items shown.")

    def on_apply(self):
        user = self.user_var.get().strip()
        code = self.code_var.get().strip()
        qty = self._parse_int(self.qty_var.get())
        box = self._parse_int(self.box_var.get(), allow_none=True)

        if not code or qty is None or qty <= 0:
            messagebox.showerror("Input error", "Enter a Part# and a positive Qty.")
            return

        try:
            if self.mode_var.get() == "ADD":
                if not confirm_action(self, "ADD", code, qty, box):
                    self.msg_var.set("Add cancelled.")
                    return
                msg = add_mode(code, qty, box, user)
            else:
                if not confirm_action(self, "REMOVE", code, qty, box):
                    self.msg_var.set("Remove cancelled.")
                    return
                msg = remove_mode(code, qty, user, box=box)

            if msg == "Code not found." or msg.startswith("Multiple boxes exist") or msg.startswith("No existing row"):
                messagebox.showwarning("Notice", msg)
            elif msg.startswith("Cannot remove") or msg.endswith("Please enter Box#."):
                messagebox.showerror("Error", msg)
            else:
                messagebox.showinfo("Done", msg)

            self.msg_var.set(msg)
            self.refresh_table()
        except PermissionError:
            messagebox.showerror("Locked", "Close the Excel file and try again.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def on_normalize_all(self):
        try:
            changed = normalize_all(self.user_var.get().strip())
            self.msg_var.set(f"Normalized {changed} item(s).")
            self.refresh_table()
            if changed:
                messagebox.showinfo("Normalization complete", f"Normalized {changed} item(s).")
        except PermissionError:
            messagebox.showerror("Locked", "Close the Excel file and try again.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def open_excel(self):
        try:
            path = EXCEL_PATH.resolve()
            import os, platform, subprocess
            if platform.system() == "Windows":
                os.startfile(str(path))
            elif platform.system() == "Darwin":
                subprocess.run(["open", str(path)])
            else:
                subprocess.run(["xdg-open", str(path)])
        except Exception as e:
            messagebox.showerror("Open failed", str(e))

if __name__ == "__main__":
    app = App()
    app.mainloop()
